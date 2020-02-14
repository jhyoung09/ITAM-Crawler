##################################################
#   James Hunter Young
#   
#   ITAM Data Compile
#   
#   this code reads from a single ITAM AssetInventory Sheet
#   then places the values into another excel document
#   
#   
#   
#   
##################################################


#   ipmorts
import openpyxl

#   global variables
preWB = openpyxl.load_workbook('AssetInventory-PRE.xlsx')
preSheet = preWB['Computers']

postWB = openpyxl.load_workbook('AssetInventory-POST.xlsx')
postSheet = postWB['Computers']

itamReport = openpyxl.load_workbook('pythonCrawl_Master.xlsx')
itamSheet = itamReport['Asset to Location Tracking All']

def getData(sheetObj):
    assetData = []
    print('Opening workbook...')
    print('... grabbing data...')
    for row in range(2, sheetObj.max_row + 1):
        PCN = sheetObj.cell(row=row, column=2).value
        if PCN is None:
            # ignore those rows which have an empty PCN (assuming that it must be present)
            continue

        if PCN == "n/a":
            # special case to deal with integer comparison where PCN was set to n/a
            PCN = 0
        deviceType = sheetObj.cell(row=row, column=3).value
        deviceSN = sheetObj.cell(row=row, column=1).value
        userID = sheetObj.cell(row=row, column=6).value

        assetData.append([PCN, deviceType, deviceSN, userID])
    return assetData

def write_data(finalData):
    print('writing data...')

    for rowNum in range(3,itamSheet.max_row):   #   skipping the first 2 rows because of headers
        for colNum in range(1,10):
            itamSheet.cell(row=rowNum, column=colNum).value = finalData
    
    print('data written in sheet... saving workbook...')
    
    itamReport.save('pythonCrawl_master.xlsx')
    
    print('workbook saved...')

def gather_data(preData, postData):
    preData.sort()
    postDataSort = sorted(postData)
    finalData = []
    preIndex = 0
    while preIndex < len(preData):
        foundMatch = False
        for postItem in postDataSort:
            if preData[preIndex][0] == postItem[0]:
                finalData.append(preData[preIndex] + postItem)
                foundMatch = True
                postDataSort.remove(postItem)

        if not foundMatch:
            finalData.append(preData[preIndex] + ["", "", "", ""])
        preIndex += 1

    for postItem in postDataSort:
        finalData.append(["", "", "", ""] + postItem)

    print("This is data that has been correlated together if possible")
    for item in finalData:
        print(item)


def main():
    preAssetData = getData(preSheet)
    postAssetData = getData(postSheet)
    finalAssetData = #  can't figure out where to corolate this to
    gather_data(preAssetData, postAssetData)
    #print(preAssetData)
    #print(postAssetData)
    write_data(finalAssetData)
    print('...DONE!!')

main()
